"""OpenTRACK: rebuild an Excel track definition and export NumPy data."""

from __future__ import annotations

import argparse
import contextlib
import math
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TextIO
import numpy as np
from openpyxl import load_workbook
from track_data import track_data
# 設定物件
base_dir = Path(__file__).resolve().parent
track_info = track_data()


show_plt = track_info.show_plt
max_speed = 110/3.6
radius_limit = 100.0
mu_y = 1.7

@dataclass
class Segment:
    # One row from the Shape sheet, normalized into geometry-friendly fields.
    section: int
    kind: str
    start: tuple[float, float]
    end: tuple[float, float]
    length: float
    center: tuple[float, float] | None = None
    radius: float | None = None
    direction: str = ""


class Tee:
    # Mirror stdout to both the terminal and the log file.
    def __init__(self, *streams: TextIO):
        self.streams = streams

    def write(self, text: str) -> int:
        for stream in self.streams:
            stream.write(text)
        return len(text)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass
class OutputPaths:
    npz: Path
    images: list[Path]
    top_radius_distribution_length: list[dict[str, float | int | str]]
    top_radius_distribution_time: list[dict[str, float | int | str]]


def read_workbook(path: Path) -> tuple[dict[str, str], list[Segment], dict[str, np.ndarray]]:
    wb = load_workbook(path, data_only=True, read_only=True)

    # Track metadata is stored in fixed rows of the Info sheet.
    info_values = [row[1] for row in wb["Info"].iter_rows(min_row=1, max_row=7, values_only=True)]
    keys = ("name", "country", "city", "type", "config", "direction", "mirror")
    info = {key: str(value) for key, value in zip(keys, info_values)}

    # Shape rows define the centerline as LINE and ARC sections.
    shape = wb["Shape"]
    headers = {str(cell.value): index for index, cell in enumerate(next(shape.iter_rows()))}
    segments: list[Segment] = []
    for row_number, row in enumerate(shape.iter_rows(min_row=2, values_only=True), start=1):
        kind = str(row[headers["Type"]] or "").strip().upper()
        if not kind:
            continue
        if kind not in {"LINE", "ARC"}:
            print(f"Warning: skipped unknown segment type in row {row_number + 1}: {kind}")
            continue
        get = lambda name: row[headers[name]]
        segment = Segment(
            section=row_number,
            kind=kind,
            start=(float(get("Start_X")), float(get("Start_Y"))),
            end=(float(get("End_X")), float(get("End_Y"))),
            length=float(get("length")),
        )
        if kind == "ARC":
            # ARC sections need circle geometry in addition to start/end points.
            segment.center = (float(get("Center_X")), float(get("Center_Y")))
            segment.radius = float(get("Radius"))
            segment.direction = str(get("Direction") or "").strip()
        segments.append(segment)

    # These sheets are section-based tables; build_track expands them to every mesh point.
    sheets = {
        "elevation": "Elevation",
        "bank": "Banking",
        "factor_grip": "Grip Factors",
        "sector": "Sectors",
    }
    tables: dict[str, np.ndarray] = {}
    for key, sheet_name in sheets.items():
        values = [row[1] for row in wb[sheet_name].iter_rows(min_row=2, values_only=True) if row[0] is not None]
        tables[key] = np.asarray(values, dtype=float)
    wb.close()
    return info, segments, tables


def reconstruct_geometry(segments: list[Segment], ds: float) -> dict[str, np.ndarray]:
    if not segments:
        raise ValueError("Shape sheet contains no valid segments.")
    # Start from the first segment start point and append generated mesh points.
    x, y = segments[0].start
    distance = 0.0
    xs: list[float] = []
    ys: list[float] = []
    ss: list[float] = []
    kappas: list[float] = []
    sections: list[int] = []
    kinds: list[str] = []

    for segment in segments:
        if segment.kind == "LINE":
            # Split straight sections into nearly uniform steps no larger than ds.
            dx = segment.end[0] - segment.start[0]
            dy = segment.end[1] - segment.start[1]
            length = math.hypot(dx, dy)
            if length == 0:
                continue
            count = max(1, math.ceil(length / ds))
            step = length / count
            ux, uy = dx / length, dy / length
            for _ in range(count):
                x += ux * step
                y += uy * step
                distance += step
                xs.append(x); ys.append(y); ss.append(distance); kappas.append(0.0)
                sections.append(segment.section); kinds.append("LINE")
        else:
            # Walk along the arc angle; clockwise arcs use negative curvature.
            assert segment.center is not None and segment.radius is not None
            cx, cy = segment.center
            theta = math.atan2(segment.start[1] - cy, segment.start[0] - cx)
            sign = -1.0 if segment.direction.lower().startswith("clock") else 1.0
            count = max(1, math.ceil(segment.length / ds))
            dtheta = (segment.length / segment.radius) / count
            for _ in range(count):
                theta += sign * dtheta
                x = cx + segment.radius * math.cos(theta)
                y = cy + segment.radius * math.sin(theta)
                distance += segment.radius * dtheta
                xs.append(x); ys.append(y); ss.append(distance); kappas.append(sign / segment.radius)
                sections.append(segment.section); kinds.append("LEFT" if sign > 0 else "RIGHT")

    return {
        "X": np.asarray(xs), "Y": np.asarray(ys), "x": np.asarray(ss),
        "kappa_raw": np.asarray(kappas), "section": np.asarray(sections, dtype=np.int32),
        "segment_type": np.asarray(kinds),
    }


def gaussian_smooth(values: np.ndarray, window: int = 80) -> np.ndarray:
    # Edge padding keeps the output length identical to the input length.
    radius = window // 2
    sigma = max(window / 6.0, np.finfo(float).eps)
    positions = np.arange(-radius, radius + 1, dtype=float)
    kernel = np.exp(-0.5 * (positions / sigma) ** 2)
    kernel /= kernel.sum()
    padded = np.pad(values, radius, mode="edge")
    return np.convolve(padded, kernel, mode="valid")


def peak_prominence(values: np.ndarray, index: int) -> float:
    # Estimate local peak prominence without depending on SciPy.
    peak = values[index]
    left_min = peak
    for i in range(index - 1, -1, -1):
        if values[i] > peak:
            break
        left_min = min(left_min, values[i])
    right_min = peak
    for i in range(index + 1, len(values)):
        if values[i] > peak:
            break
        right_min = min(right_min, values[i])
    return peak - max(left_min, right_min)


def find_peaks(values: np.ndarray, minimum_prominence: float, minimum_distance: int) -> np.ndarray:
    # Pick strong local maxima first, then enforce a minimum index spacing.
    candidates = [
        i for i in range(1, len(values) - 1)
        if values[i] > values[i - 1] and values[i] >= values[i + 1]
        and peak_prominence(values, i) >= minimum_prominence
    ]
    selected: list[int] = []
    for index in sorted(candidates, key=lambda i: values[i], reverse=True):
        if all(abs(index - other) >= minimum_distance for other in selected):
            selected.append(index)
    return np.asarray(sorted(selected), dtype=np.int32)


def build_track(segments: list[Segment], tables: dict[str, np.ndarray], ds: float) -> dict[str, np.ndarray]:
    track = reconstruct_geometry(segments, ds)
    raw = track["kappa_raw"]
    # Apex detection is based on smoothed curvature, not the raw stepped curvature.
    kappa = gaussian_smooth(raw)
    prominence = float(np.percentile(np.abs(kappa), 70) * 0.20)
    distance = max(1, round(8.0 / ds))
    left = find_peaks(kappa, prominence, distance)
    right = find_peaks(-kappa, prominence, distance)
    apex = np.asarray(sorted(np.concatenate((left, right)).tolist()), dtype=np.int32)

    section_indices = track["section"] - 1
    # Expand section-level values so every generated mesh point has matching data.
    for key, values in tables.items():
        if values.size == 0:
            raise ValueError(f"Input table {key!r} is empty.")
        track[key] = values[np.clip(section_indices, 0, len(values) - 1)]
    s = track["x"]
    elevation = track.pop("elevation")
    # Convert elevation slope into inclination angle and smooth small point-to-point noise.
    inclination = -np.degrees(np.diff(elevation) / np.diff(s))
    inclination = np.append(inclination, inclination[-1] if inclination.size else 0.0)
    inclination = np.convolve(np.pad(inclination, 2, mode="edge"), np.ones(5) / 5, mode="valid")

    track.update({
        "Z": elevation, "r": kappa, "incl": inclination,
        "dx": np.gradient(s), "n": np.asarray(len(s), dtype=np.int64),
        "apex_index": apex, "apex_s": s[apex],
        "apex_type": np.sign(kappa[apex]).astype(np.int8),
        "apex_strength": np.abs(kappa[apex]),
        "apex_mask": np.isin(np.arange(len(s)), apex),
        "arrow": np.asarray(0, dtype=np.int8),
    })
    return track


def ascii_map(track: dict[str, np.ndarray], width: int = 66) -> str:
    # Lightweight terminal preview for the log file.
    x, y = track["X"], track["Y"]
    span = float(np.ptp(x)) or 1.0
    xx = np.rint(x / span * width).astype(int)
    yy = np.rint(y / (15 / 8) / span * width).astype(int)
    xx -= xx.min(); yy = -yy; yy -= yy.min()
    points = set(zip(xx.tolist(), yy.tolist()))
    return "\n".join("".join("o" if (col, row) in points else " " for col in range(xx.max() + 1)).rstrip()
                     for row in range(yy.max() + 1))


def build_ranked_radius_distribution(
    values: np.ndarray,
    edges: np.ndarray,
    centers: np.ndarray,
    metric_key: str,
) -> list[dict[str, float | int | str]]:
    ranked_indices = np.argsort(-values, kind="stable")
    return [
        {
            "rank": rank,
            "radius_range_m": f"{edges[index]:.0f}-{edges[index + 1]:.0f}",
            "radius_center_m": float(centers[index]),
            metric_key: float(values[index]),
        }
        for rank, index in enumerate(ranked_indices, start=1)
    ]


def save_pngs(
    track: dict[str, np.ndarray],
    output_dir: Path,
    stem: str,
) -> tuple[list[Path], list[dict[str, float | int | str]], list[dict[str, float | int | str]]]:
    # Keep Matplotlib cache inside the output folder to avoid user-profile permission issues.
    cache_dir = output_dir / ".matplotlib-cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(cache_dir))
    import matplotlib
    if not show_plt:
        matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.collections import LineCollection
    from matplotlib.colors import TwoSlopeNorm

    x, y = track["X"], track["Y"]
    raw_curvature = track["kappa_raw"]
    curvature = track["r"]
    curvature_eps = 1e-9
    raw_radius = np.full_like(raw_curvature, np.nan, dtype=float)
    radius = np.full_like(curvature, np.nan, dtype=float)
    raw_curve_mask = np.abs(raw_curvature) > curvature_eps
    curve_mask = np.abs(curvature) > curvature_eps
    raw_radius[raw_curve_mask] = 1.0 / np.abs(raw_curvature[raw_curve_mask])
    radius[curve_mask] = 1.0 / np.abs(curvature[curve_mask])
    apex = track["apex_index"]
    left_apex = apex[track["apex_type"] >= 0]
    right_apex = apex[track["apex_type"] < 0]

    figures: list[plt.Figure] = []
    image_paths: list[Path] = []

    def save_current(fig: plt.Figure, suffix: str) -> None:
        path = output_dir / f"{stem}_{suffix}.png"
        figures.append(fig)
        image_paths.append(path)
        plt.figure(fig.number)
        plt.savefig(path, format="png", dpi=180)

    # XY map: color the track by smoothed curvature and add apex markers.
    fig, ax = plt.subplots(figsize=(10, 8), constrained_layout=True)
    points = np.column_stack((x, y))
    segments = np.stack((points[:-1], points[1:]), axis=1)
    limit = float(np.max(np.abs(curvature))) or 1.0
    norm = TwoSlopeNorm(vmin=-limit, vcenter=0.0, vmax=limit)
    line_values = (curvature[:-1] + curvature[1:]) / 2.0
    colored_track = LineCollection(segments, cmap="rainbow", norm=norm, linewidth=3.0)
    colored_track.set_array(line_values)
    ax.add_collection(colored_track)
    ax.autoscale()
    ax.set_aspect("equal", adjustable="box")
    ax.set_title("XY Map")
    ax.set_xlabel("X [m]")
    ax.set_ylabel("Y [m]")
    ax.grid(True)
    cbar = fig.colorbar(colored_track, ax=ax)
    cbar.set_label("Curvature [1/m]")
    # 方便觀察曲率變化 先隱藏apex
    #ax.scatter(x[left_apex], y[left_apex], s=32, c="#ffb000", edgecolors="black", linewidths=0.5, label="Left apex", zorder=3)
    #ax.scatter(x[right_apex], y[right_apex], s=32, c="#00a6d6", edgecolors="black", linewidths=0.5, label="Right apex", zorder=3)
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles, labels, loc="best")
    save_current(fig, "xy_map")

    # Curvature comparison: apex points are marked only on the smoothed curve.
    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    ax.plot(track["x"], raw_curvature, color="#2ca02c", linewidth=1, label="Raw curvature")
    ax.plot(track["x"], curvature, color="#1f77b4", linewidth=1, label="Gaussian smoothed")
    ax.scatter(
        track["apex_s"],
        curvature[apex],
        s=22,
        c="#c13636",
        linewidths=0.35,
        label="Apex after smoothing",
        zorder=3,
    )
    ax.set_title("Curvature vs Distance")
    ax.set_xlabel("Distance [m]")
    ax.set_ylabel("Curvature [1/m]")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best")
    save_current(fig, "radius")

    # FFT spectrum of smoothed curvature over distance.
    ds_fft = float(np.mean(track["dx"])) if len(track["dx"]) else 1.0
    fft_val = np.fft.fft(curvature)
    freq = np.fft.fftfreq(len(curvature), d=ds_fft)
    amp = np.abs(fft_val)
    positive_freq_mask = freq > 0
    positive_freq = freq[positive_freq_mask]
    positive_amp = amp[positive_freq_mask]

    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    freq_limit = 0.5
    fft_plot_mask = positive_freq <= freq_limit
    ax.plot(positive_freq[fft_plot_mask], positive_amp[fft_plot_mask], color="#1f77b4", linewidth=1.2)
    ax.set_xlim(0.0, freq_limit)
    ax.set_title("Curvature Spectrum (FFT)")
    ax.set_xlabel("Spatial Frequency [1/m]")
    ax.set_ylabel("Amplitude")
    ax.grid(True, alpha=0.3)
    save_current(fig, "curvature_fft_frequency")

    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    wavelength = 1.0 / positive_freq
    order = np.argsort(wavelength)
    ax.plot(wavelength[order], positive_amp[order], color="#1f77b4", linewidth=1.2)
    ax.set_xscale("log")
    ax.set_title("Track Curvature Spectrum")
    ax.set_xlabel("Spatial Wavelength [m]")
    ax.set_ylabel("Amplitude")
    ax.grid(True, which="both", alpha=0.3)
    save_current(fig, "curvature_fft_wavelength")

    # Radius time distribution: each point contributes dx / v to a 1 m radius bin.
    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    finite_radius_mask = np.isfinite(radius) & (radius <= radius_limit)
    finite_radius = radius[finite_radius_mask]
    finite_dx = track["dx"][finite_radius_mask]
    muy = float(mu_y)
    corner_speed = np.sqrt(np.maximum(muy * finite_radius, 0.0))
    limited_speed = np.minimum(corner_speed, max_speed)
    finite_dt = np.divide(
        finite_dx,
        limited_speed,
        out=np.full_like(finite_dx, np.nan, dtype=float),
        where=limited_speed > 1e-9,
    )
    valid_time_mask = np.isfinite(finite_dt)
    bins = np.arange(0.0, radius_limit + 1.0, 1.0)
    length_by_radius, _ = np.histogram(finite_radius, bins=bins, weights=finite_dx)
    time_by_radius, edges = np.histogram(finite_radius[valid_time_mask], bins=bins, weights=finite_dt[valid_time_mask])
    centers = (edges[:-1] + edges[1:]) / 2.0
    widths = np.diff(edges)
    top_time_indices = np.argsort(-time_by_radius, kind="stable")[:5]
    top_length_indices = np.argsort(-length_by_radius, kind="stable")[:5]
    top_index_set = set(top_time_indices.tolist())
    bar_colors = ["#d62728" if index in top_index_set else "#1f77b4" for index in range(len(time_by_radius))]
    ax.bar(centers, time_by_radius, width=widths, align="center", color=bar_colors, linewidth=0.3)
    top_radius_distribution_time = build_ranked_radius_distribution(
        time_by_radius,
        edges,
        centers,
        "accumulated_time_s",
    )
    top_radius_distribution_length = build_ranked_radius_distribution(
        length_by_radius,
        edges,
        centers,
        "accumulated_length_m",
    )
    ax.set_title("Radius Time Distribution (0-100 m)")
    ax.set_xlabel("Radius [m]")
    ax.set_ylabel("Accumulated time [s]")
    ax.grid(True, axis="y", alpha=0.3)
    ax.legend(
        handles=[
            plt.Rectangle((0, 0), 1, 1, color="#d62728", label="Top 5 time bins"),
            plt.Rectangle((0, 0), 1, 1, color="#1f77b4", label="Other radius bins"),
        ],
        loc="best",
    )
    save_current(fig, "radius_time_distribution")

    fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
    bar_colors = ["#d62728" if index in set(top_length_indices.tolist()) else "#1f77b4" for index in range(len(length_by_radius))]
    ax.bar(centers, length_by_radius, width=widths, align="center", color=bar_colors, linewidth=0.3)
    ax.set_title("Radius Length Distribution (0-100 m)")
    ax.set_xlabel("Radius [m]")
    ax.set_ylabel("Accumulated length [m]")
    ax.grid(True, axis="y", alpha=0.3)
    ax.legend(
        handles=[
            plt.Rectangle((0, 0), 1, 1, color="#d62728", label="Top 5 length bins"),
            plt.Rectangle((0, 0), 1, 1, color="#1f77b4", label="Other radius bins"),
        ],
        loc="best",
    )
    save_current(fig, "radius_length_distribution")

    # Remaining plots are section data expanded onto the generated distance vector.
    scalar_plots = [
        ("elevation", "Elevation", track["x"], track["Z"], "Distance [m]", "Elevation [m]"),
        ("inclination", "Inclination", track["x"], track["incl"], "Distance [m]", "Inclination [deg]"),
        ("banking", "Banking", track["x"], track["bank"], "Distance [m]", "Banking [deg]"),
        ("grip_factor", "Grip Factor", track["x"], track["factor_grip"], "Distance [m]", "Grip factor [-]"),
    ]
    for suffix, title, xx, yy, xlabel, ylabel in scalar_plots:
        fig, ax = plt.subplots(figsize=(12, 5), constrained_layout=True)
        ax.plot(xx, yy, color="#1f77b4", linewidth=1.4)
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True, alpha=0.3)
        save_current(fig, suffix)

    # Display figures only when requested by track_info; otherwise release memory.
    if show_plt:
        plt.show()  # Optional interactive display.
    else:
        plt.close("all")  # Avoid memory growth when running batch jobs.
    for fig in figures:
        plt.close(fig)
    return image_paths, top_radius_distribution_length, top_radius_distribution_time

def run(input_path: Path, output_dir: Path, mesh_size: float) -> OutputPaths:
    info, segments, tables = read_workbook(input_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"OpenTRACK"
    log_path = output_dir / f"{stem}.log"
    with log_path.open("w", encoding="utf-8") as log, contextlib.redirect_stdout(Tee(sys.stdout, log)):
        print("OpenTRACK Python")
        print("=" * 66)
        print(f"Input:         {input_path}")
        for label, key in (("Name", "name"), ("City", "city"), ("Country", "country"),
                           ("Type", "type"), ("Configuration", "config"),
                           ("Direction", "direction"), ("Mirror", "mirror")):
            print(f"{label + ':':15}{info[key]}")
        print(f"Generated:     {datetime.now().isoformat(timespec='seconds')}")
        print("=" * 66)
        track = build_track(segments, tables, mesh_size)
        print(f"Track points:  {int(track['n'])}")
        print(f"Track length:  {track['x'][-1]:.3f} m")
        print(f"Mesh size:     {mesh_size:g} m (maximum target spacing)")
        print(f"Apex count:    {len(track['apex_index'])}")
        nonzero_curvature = np.abs(track["r"]) > 1e-9
        min_radius = float(np.min(1.0 / np.abs(track["r"][nonzero_curvature]))) if np.any(nonzero_curvature) else math.inf
        radius = np.full_like(track["r"], np.nan, dtype=float)
        radius[nonzero_curvature] = 1.0 / np.abs(track["r"][nonzero_curvature])
        turn_radius_mask = np.isfinite(radius) & (radius <= 100.0)
        average_radius = (
            float(np.average(radius[turn_radius_mask], weights=track["dx"][turn_radius_mask]))
            if np.any(turn_radius_mask)
            else math.inf
        )
        print(f"Minimum radius:{min_radius:.3f} m")
        print("\nMap:\n" + ascii_map(track))

        npz_path = output_dir / f"{stem}.npz"
        info_arrays = {f"info_{key}": np.asarray(value) for key, value in info.items()}
        np.savez_compressed(npz_path, **track, **info_arrays)
        image_paths, top_radius_distribution_length, top_radius_distribution_time = save_pngs(track, output_dir, stem)
        track_info.Track_length = float(track["x"][-1])
        track_info.Average_radius = average_radius
        track_info.Name = info["name"]
        track_info.Country = info["country"]
        track_info.Track_points = int(track["n"])
        track_info.Mesh_size = float(mesh_size)
        track_info.Apex_count = int(len(track["apex_index"]))
        track_info.track_Minimum_radius = min_radius
        track_info.Radius_distribution_length = top_radius_distribution_length
        track_info.Radius_distribution_time = top_radius_distribution_time
        track_info.save_json(base_dir / "track.json")
        print("Updated track.json with open track")
        print(f"\nNumPy track saved: {npz_path}")
        print("Visualizations saved:")
        for image_path in image_paths:
            print(f"  {image_path}")
    return OutputPaths(
        npz=npz_path,
        images=image_paths,
        top_radius_distribution_length=top_radius_distribution_length,
        top_radius_distribution_time=top_radius_distribution_time,
    )



parser = argparse.ArgumentParser(description="Convert an OpenTRACK Excel map to NumPy NPZ data.")
parser.add_argument("input", nargs="?", type=Path, default=Path(__file__).with_name("input_map.xlsx"))
parser.add_argument("--mesh-size", type=float, default=0.1, help="Maximum point spacing in metres (default: 0.5)")
parser.add_argument("--output", type=Path, default=Path(__file__).with_name("OpenTRACK Tracks"))
args = parser.parse_args()
if args.mesh_size <= 0:
    parser.error("--mesh-size must be greater than zero")
run(args.input.resolve(), args.output.resolve(), args.mesh_size)
